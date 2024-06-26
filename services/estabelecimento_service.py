from typing import List, Dict
from .default_service import DefaultService
from repositories.estabelecimento_repository import EstabelecimentoRepository
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def get_situacao(situacao_cadastral):

  match situacao_cadastral:
    case 1:
      situacao = "Nula"
    case 2:
      situacao = "Ativa"
    case 3:
      situacao = "Suspensa"
    case 4:
      situacao = "Inapta"
    case 8:
      situacao = "Baixada"

  return situacao

def get_porte(porte_empresa):
  match porte_empresa:
    case 1:
      porte = "Não informado"
    case 2:
      porte = "Micro empresa"
    case 3:
      porte = "Empresa de pequeno porte"
    case 5:
      porte = "Demais"
  
  return porte


def criar_pdf(nome_arquivo, estabelecimentos):

  pdf = FPDF()
  #pdf.set_auto_page_break(auto=True, margin=15)
  pdf.set_margins(20, 20, 20)

  pdf.add_page()
  pdf.set_font("Helvetica", "B", size=16)

  pdf.cell(0, 10, "Estabelecimentos", ln=1, align='C')
  pdf.ln(7)

  pdf.set_font("Helvetica", size=11)

  estabelecimentos_na_pagina = 0

  for estabelecimento in estabelecimentos:
    estabelecimentos_na_pagina += 1

    if (estabelecimento.nome_fantasia):
      pdf.cell(0, 10, f"Nome Fantasia: {estabelecimento.nome_fantasia}", ln=1)

    pdf.cell(0, 10, f"Razão Social: {estabelecimento.empresa.razao_social}", ln=1)
    pdf.cell(0, 10, f"CNPJ Básico: {estabelecimento.cnpj_basico}", ln=1)
    pdf.multi_cell(0, 10, f"CNAE: {estabelecimento.empresa.cnae.id} - {estabelecimento.empresa.cnae.descricao}", align='J')

    situacao = get_situacao(estabelecimento.situacao_cadastral)

    pdf.cell(0, 10, f"Situação Cadastral: {estabelecimento.situacao_cadastral} - {situacao}", ln=1)

    porte = get_porte(estabelecimento.empresa.porte)

    pdf.cell(0, 10, f"Porte: {estabelecimento.empresa.porte} - {porte}", ln=1)
    pdf.cell(0, 10, f"Natureza Jurídica: {estabelecimento.empresa.natureza_juridica.descricao}", ln=1)
    pdf.cell(0, 10, f"Rua: {estabelecimento.endereco.logradouro}", ln=1)
    pdf.cell(0, 10, f"Número: {estabelecimento.endereco.numero}", ln=1)
    pdf.cell(0, 10, f"Bairro: {estabelecimento.endereco.bairro}", ln=1)
    pdf.cell(0, 10, f"CEP: {estabelecimento.endereco.cep}", ln=1)
    pdf.cell(0, 10, f"Cidade: {estabelecimento.endereco.municipio.descricao}", ln=1)
    pdf.ln(7)

    if estabelecimentos_na_pagina % 2 == 0:
      pdf.add_page()


  pdf.output(nome_arquivo)

  print(f"PDF '{nome_arquivo}' criado com sucesso.")


def criar_xlsx(nome_arquivo, estabelecimentos):

  wb = Workbook()
  sheet = wb.active

  sheet.append([
    "Nome Fantasia",
    "Razão Social",
    "CNPJ Básico",
    "CNAE",
    "Descrição",
    "Situação Cadastral",
    "Porte",
    "Natureza Jurídica",
    "Endereço",
    "Cidade"
  ])

  for cell in sheet[1]:
    cell.font = Font(name='Arial', size=12, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

  for estabelecimento in estabelecimentos:
    situacao = get_situacao(estabelecimento.situacao_cadastral)
    porte = get_porte(estabelecimento.empresa.porte)
    sheet.append([
      estabelecimento.nome_fantasia,
      estabelecimento.empresa.razao_social,
      estabelecimento.cnpj_basico,
      estabelecimento.empresa.cnae.id,
      estabelecimento.empresa.cnae.descricao,
      f"{estabelecimento.situacao_cadastral} - {situacao}",
      f"{estabelecimento.empresa.porte} - {porte}",
      estabelecimento.empresa.natureza_juridica.descricao,
      f"RUA {estabelecimento.endereco.logradouro}, {estabelecimento.endereco.numero}, {estabelecimento.endereco.bairro}, CEP {estabelecimento.endereco.cep}",
      estabelecimento.endereco.municipio.descricao
    ])

  for row in sheet.iter_rows(min_row=2, max_row=len(estabelecimentos) + 1):
    for cell in row:
      cell.font = Font(name='Arial', size=11)
      cell.alignment = Alignment(horizontal='left', vertical='center')

  for col in sheet.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
      try:
        if len(str(cell.value)) > max_length:
          max_length = len(cell.value) + 10
      except:
        pass
    adjusted_width = (max_length + 2) * 1.2
    sheet.column_dimensions[column].width = adjusted_width

  wb.save(nome_arquivo)

  print(f"XLSX '{nome_arquivo}' criado com sucesso.")


class EstabelecimentoService(DefaultService):
  def __init__(self):
    super().__init__(EstabelecimentoRepository)
    
  def get_estab_by_id_with_empresa(self, id: int) -> Dict:
    estabelecimento = self.repository.get_by_id(id)  # Supondo que este método retorna um único estabelecimento
    if not estabelecimento:
        return None  # Ou manipule de acordo com a necessidade, talvez levantando uma exceção
    
    filters = {
      'cnae': estabelecimento.empresa.cnae_principal_id,
    }

    estabelecimentos_relacionados = self.repository.get_filtered_estab(filters)
    for empresa in estabelecimentos_relacionados:
      if empresa.id == estabelecimento.empresa.id:
        estabelecimentos_relacionados.remove(empresa)

    # Serializar o estabelecimento incluindo os detalhes da empresa relacionada
    return self._serialize(estabelecimento, estabelecimentos_relacionados, include_empresa=True, include_socios=True, include_address=True, include_relacionadas=True)



  def get_all(self, page: int) -> List[Dict]:
    estabelecimentos = self.repository.get_all(page)
    return [self._serialize_simple(estabelecimento) for estabelecimento in estabelecimentos]

  def get_by_id(self, id: int) -> Dict:
    estabelecimento = self.repository.get_by_id(id)
    return self._serialize(estabelecimento) 
  
  def get_by_cnpj(self, cnpj: int) -> List[Dict]:
    estabelecimentos = self.repository.get_by_cnae(cnpj)
    return [self._serialize(estabelecimento) for estabelecimento in estabelecimentos]
  
  def get_filtered(self, filters: Dict) -> List[Dict]:
    page = self.repository.get_filtered(filters)

    estabelecimentos = page['estabelecimentos']
    total = page['total']

    criar_pdf("./download/estabelecimentos.pdf", estabelecimentos)
    criar_xlsx("./download/estabelecimentos.xlsx", estabelecimentos)

    return { 'estabelecimentos': [self._serialize_simple(estabelecimento) for estabelecimento in estabelecimentos], 'total': total}
  
  def _serialize_empresa(self,estabelecimento_relacionado, include_address:bool=False):
    
      serialized_data = {
        # 'id_estab_empresa_id':
        'id_empresa': estabelecimento_relacionado.empresa_id,
        'id_estabelecimento': estabelecimento_relacionado.id,
        'cnpj_basico': estabelecimento_relacionado.cnpj_basico,
        'porte': estabelecimento_relacionado.empresa.porte,
        'razao_social': estabelecimento_relacionado.empresa.razao_social,
        'natureza_juridica_id': estabelecimento_relacionado.empresa.natureza_juridica_id,
        'capital_social': estabelecimento_relacionado.empresa.capital_social,
        'cnae_principal_id': estabelecimento_relacionado.empresa.cnae_principal_id,
        # 'endereco':estabelcimento_relacionado.endereco if estabelcimento_relacionado.endereco else None,
        
      }
      if include_address and estabelecimento_relacionado.endereco:
          serialized_data['endereco'] = {
          'logradouro': estabelecimento_relacionado.endereco.logradouro,
          'numero': estabelecimento_relacionado.endereco.numero,
          'bairro': estabelecimento_relacionado.endereco.bairro,
          'cidade':estabelecimento_relacionado.endereco.municipio.descricao if estabelecimento_relacionado.endereco.municipio else '',
          'cep':estabelecimento_relacionado.endereco.cep,
          'municipio': estabelecimento_relacionado.endereco.municipio.descricao if estabelecimento_relacionado.endereco.municipio else ''
          # ... outros campos do endereço
        }
        
      return serialized_data
  
  def _serialize_socio(self,socio_empresa):
    return {
        'id': socio_empresa.socio.id,
        'cpf_cnpj': socio_empresa.socio.cpf_cnpj,
        'nome_socio': socio_empresa.socio.nome_socio,
        'representante_legal': socio_empresa.socio.representante_legal,
        'nome_representante_legal': socio_empresa.socio.nome_representante_legal,
        # ... outros campos do sócio
    }

  def _serialize(self, estabelecimento, estabelecimentos_relacionados, include_address:bool=False, include_empresa: bool = False, include_socios:bool = False, include_relacionadas:bool = False) -> Dict:
    # Esta função assume que sua entidade `Estabelecimento` é um modelo SQLAlchemy
    # e converte para um dicionário. Você pode precisar ajustar isso
    # para se adequar à estrutura exata de sua entidade `Estabelecimento`.
    serialized_data = {
    	'id_estabelecimento': estabelecimento.id,
      'cnae': estabelecimento.empresa.cnae.descricao,
      'cnpj_basico': estabelecimento.cnpj_basico,
      'cnpj_ordem': estabelecimento.cnpj_ordem,
      'identificador_matriz_filial': estabelecimento.identificador_matriz_filial,
      'nome_fantasia': estabelecimento.nome_fantasia,
      'data_inicio_atividade': estabelecimento.data_inicio_atividade,
      'endereco_id': estabelecimento.endereco_id,
      'situacao_cadastral': estabelecimento.situacao_cadastral,
      'nome_fantasia':estabelecimento.nome_fantasia,
    }
    if include_empresa and estabelecimento.empresa:
        empresa_data = {
            'id_empresa': estabelecimento.empresa.id,
            'razao_social': estabelecimento.empresa.razao_social,
            'capital_social': str(estabelecimento.empresa.capital_social),
            'porte': estabelecimento.empresa.porte,
            'natureza_juridica_descricao': estabelecimento.empresa.natureza_juridica.descricao if estabelecimento.empresa.natureza_juridica else None,
            'cnae_descricao': estabelecimento.empresa.cnae.descricao if estabelecimento.empresa.cnae else None,
        }
            # Aqui você pode adicionar mais campos conforme necessário
        if include_socios:
            empresa_data['socios'] = [
                self._serialize_socio(socio) for socio in estabelecimento.empresa.socio_empresas[:10]
            ]

        if include_relacionadas:
          empresa_data['estabs_relacionados'] = [
              self._serialize_empresa(estabs,include_address=True) for estabs in estabelecimentos_relacionados[:5]
          ]

    serialized_data['empresa'] = empresa_data
      
    if include_address and estabelecimento.endereco:
      serialized_data['endereco'] = {
        'logradouro': estabelecimento.endereco.logradouro,
        'numero': estabelecimento.endereco.numero,
        'bairro': estabelecimento.endereco.bairro,
        'cidade':estabelecimento.endereco.municipio.descricao if estabelecimento.endereco.municipio else '',
        'cep':estabelecimento.endereco.cep,
        'municipio': estabelecimento.endereco.municipio.descricao if estabelecimento.endereco.municipio else ''
        # ... outros campos do endereço
      }
    return serialized_data
  
  def _serialize_simple(self, estabelecimento) -> Dict:
    return {
        'id': estabelecimento.id,
        'cnpj_basico': estabelecimento.cnpj_basico,
        'nome_fantasia': estabelecimento.nome_fantasia,
        'cnae_id': estabelecimento.empresa.cnae.id if estabelecimento.empresa and estabelecimento.empresa.cnae else None,
        'cnae':estabelecimento.empresa.cnae.descricao if estabelecimento.empresa and estabelecimento.empresa.cnae else None,
        'razao_social': estabelecimento.empresa.razao_social if estabelecimento.empresa else None,
        'endereco': estabelecimento.endereco.logradouro + ' - ' + estabelecimento.endereco.numero + ', ' + estabelecimento.endereco.bairro if estabelecimento.endereco else None,
        'latitude': estabelecimento.endereco.latitude if estabelecimento.endereco else None,
        'longitude': estabelecimento.endereco.longitude if estabelecimento.endereco else None,
    }
